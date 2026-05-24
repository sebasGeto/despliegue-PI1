import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render

from reportes.services import obtener_datos_reporte_asistencia
from citas.models import Caso
from usuarios.models import Usuario

ROLES_PERMITIDOS = ['administrador', 'secretaria']

COLUMNAS = ['Beneficiario', 'Caso', 'Sala', 'Fecha', 'Hora', 'Tipo atención', 'Asistió']


def _leer_filtros(request):
    filtros = {}
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sala_juridica = request.GET.get('sala_juridica')
    estudiante_id = request.GET.get('estudiante_id')
    if fecha_inicio:
        filtros['fecha_inicio'] = fecha_inicio
    if fecha_fin:
        filtros['fecha_fin'] = fecha_fin
    if sala_juridica:
        filtros['sala_juridica'] = sala_juridica
    if estudiante_id:
        filtros['estudiante_id'] = estudiante_id
    return filtros or None


def _fila(registro):
    cita = registro.cita
    beneficiario = cita.beneficiario.get_full_name() or cita.beneficiario.username
    if cita.caso:
        caso = cita.caso.codigo
        sala = cita.caso.get_sala_juridica_display()
    else:
        caso = 'Sin caso'
        sala = 'Sin caso'
    fecha = str(cita.horario.fecha)
    hora = str(cita.horario.hora_inicio)
    tipo = cita.get_tipo_atencion_display()
    asistio = 'Sí' if registro.asistio else 'No'
    return [beneficiario, caso, sala, fecha, hora, tipo, asistio]


@login_required
def exportar_reporte_pdf(request):
    if request.user.rol not in ROLES_PERMITIDOS:
        return HttpResponseForbidden('No tiene permisos para exportar reportes.')

    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    filtros = _leer_filtros(request)
    qs = obtener_datos_reporte_asistencia(filtros)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Reporte de Asistencia — Consultorio Jurídico ICESI', styles['Title']))
    elements.append(Paragraph(f'Fecha de generación: {date.today()}', styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [COLUMNAS]
    if qs.exists():
        for registro in qs:
            data.append(_fila(registro))
    else:
        data.append(['No hay registros para los filtros seleccionados'] + [''] * (len(COLUMNAS) - 1))

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
    ]))
    elements.append(tabla)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.pdf"'
    return response


@login_required
def exportar_reporte_excel(request):
    if request.user.rol not in ROLES_PERMITIDOS:
        return HttpResponseForbidden('No tiene permisos para exportar reportes.')

    from openpyxl import Workbook
    from openpyxl.styles import Font

    filtros = _leer_filtros(request)
    qs = obtener_datos_reporte_asistencia(filtros)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte Asistencia'

    ws.append(COLUMNAS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if qs.exists():
        for registro in qs:
            ws.append(_fila(registro))
    else:
        ws['A2'] = 'No hay registros para los filtros seleccionados'

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(
            20, max((len(str(cell.value or '')) for cell in col), default=0) + 2
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.xlsx"'
    return response


@login_required
def dashboard_metricas_view(request):
    """Vista simple que muestra métricas históricas y el último valor calculado.

    Permite a administradores y secretarias ver el porcentaje de asistencia
    y una gráfica histórica con los últimos registros.
    """
    if request.user.rol not in ROLES_PERMITIDOS:
        return HttpResponseForbidden('No tiene permisos para ver métricas.')

    from reportes.models import MetricasAsistencia

    filtros = _leer_filtros(request)
    registros = MetricasAsistencia.objects.all().order_by('-fecha_generacion')[:12]
    registros = list(reversed(registros))  # invertir para que el más antiguo esté primero

    labels = [r.fecha_generacion.strftime('%Y-%m-%d %H:%M') for r in registros]
    datos_asistencia = [float(r.porcentaje_asistencia) for r in registros]
    datos_inasistencia = [float(r.porcentaje_inasistencia) for r in registros]

    ultimo = MetricasAsistencia.objects.order_by('-fecha_generacion').first()

    reporte_qs = obtener_datos_reporte_asistencia(filtros)
    total_registros = reporte_qs.count()
    asistidos_count = reporte_qs.filter(asistio=True).count()
    no_asistidos_count = total_registros - asistidos_count

    return render(request, 'reportes/dashboard_metricas.html', {
        'labels': labels,
        'datos_asistencia': datos_asistencia,
        'datos_inasistencia': datos_inasistencia,
        'ultimo': ultimo,
        'filtros': filtros or {},
        'sala_choices': Caso.SALA_CHOICES,
        'estudiantes': Usuario.objects.filter(rol='estudiante').order_by('nombre_completo'),
        'total_registros': total_registros,
        'asistidos_count': asistidos_count,
        'no_asistidos_count': no_asistidos_count,
    })
